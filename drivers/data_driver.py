"""
测试数据加载器 - 支持 JSON 和 Excel 格式
"""
import json
import re
from typing import List, Dict, Any
from pathlib import Path

import openpyxl


class DataDriver:
    """测试数据驱动器 - 加载并解析 JSON/Excel 测试数据"""
    
    # 必填字段
    REQUIRED_FIELDS = ['id', 'name', 'type', 'expected_type', 'expected_value']
    
    def __init__(self, data_path: str):
        self.data_path = Path(data_path)
        if not self.data_path.exists():
            raise FileNotFoundError(f"测试数据文件不存在: {data_path}")
    
    def load(self) -> List[Dict[str, Any]]:
        """加载测试数据（自动识别格式）"""
        if self.data_path.suffix.lower() in ('.xlsx', '.xls'):
            data = self._load_excel()
        elif self.data_path.suffix.lower() == '.json':
            data = self._load_json()
        else:
            raise ValueError(f"不支持的数据格式: {self.data_path.suffix}")

        for case in data:
            missing = [f for f in self.REQUIRED_FIELDS if f not in case or not case.get(f)]
            if missing:
                raise ValueError(f"用例 {case.get('id', 'unknown')} 缺少必填字段: {missing}")
        return data
    
    def _load_json(self) -> List[Dict[str, Any]]:
        """加载 JSON 数据"""
        with open(self.data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 支持列表或字典格式
        if isinstance(data, dict):
            data = data.get('test_cases', [data])
        
        # 补充缺失字段
        for case in data:
            self._fill_missing_fields(case)
        
        return data
    
    def _load_excel(self) -> List[Dict[str, Any]]:
        """加载 Excel 数据（支持多个 Sheet）"""
        workbook = openpyxl.load_workbook(self.data_path)
        all_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 读取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else '')
            
            # 读取数据行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 跳过空行
                if not any(row):
                    continue
                
                # 构建字典
                case = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        case[headers[i]] = value if value is not None else ''
                
                # 补充缺失字段
                self._fill_missing_fields(case)
                all_data.append(case)
        
        return all_data
    
    def _fill_missing_fields(self, case: Dict[str, Any]):
        """补充缺失字段"""
        defaults = {
            'selector': '',
            'by': 'id',
            'value': '',
            'direction': 'up',
            'platform': 'All',
            'priority': 'P1',
            'description': '',
            'steps': [],
        }
        
        for key, default_value in defaults.items():
            if key not in case or not case[key]:
                case[key] = default_value
    
    @staticmethod
    def replace_placeholders(data: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """
        替换占位符 ${var_name} 为实际值
        
        Args:
            data: 测试数据字典
            context: 上下文变量字典
        
        Returns:
            替换后的字典
        """
        if not context:
            return data
        
        pattern = re.compile(r'\$\{(\w+)\}')
        
        def replace_value(value):
            if isinstance(value, str):
                def replacer(match):
                    var_name = match.group(1)
                    return str(context.get(var_name, match.group(0)))
                return pattern.sub(replacer, value)
            elif isinstance(value, dict):
                return {k: replace_value(v) for k, v in value.items()}
            elif isinstance(value, list):
                return [replace_value(item) for item in value]
            return value
        
        return {k: replace_value(v) for k, v in data.items()}
